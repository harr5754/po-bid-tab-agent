"""
Classic Excel export that fills the user's Bid Tab Template.
Preserves her layout, zebra shading, and the 20% Vendor Data Costs formulas.
"""

from __future__ import annotations
from io import BytesIO
from pathlib import Path
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.models import BidTab, calculate_apples_to_apples


TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "data" / "bid_tab_template.xlsx"


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int = 9):
    """Copy cell styles from source_row to target_row so zebra shading is preserved."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        if src.has_style:
            tgt.font = copy(src.font)
            tgt.fill = copy(src.fill)
            tgt.border = copy(src.border)
            tgt.alignment = copy(src.alignment)
            tgt.number_format = src.number_format


def build_classic_from_template(bt: BidTab) -> bytes:
    """
    Open the 3-bidder Bid Tab Template and fill it with data from the BidTab model.
    Keeps the 20% Vendor Data Costs formulas (=Dxx*0.2 etc.).
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Commercial Evaluation"]
    ws2 = wb["Technical Evaluation"]

    rfq = bt.rfq
    vendors = bt.vendors[:3]  # template supports max 3

    # ------------------------------------------------------------------
    # 1. Header block
    # ------------------------------------------------------------------
    title = f"{rfq.project_number}  {rfq.project_name}  — {rfq.rfq_number} — {rfq.equipment_tag} {rfq.equipment_description} — BID TAB"
    ws["A1"] = title
    ws2["A1"] = f"{rfq.project_number}  {rfq.project_name}  — {rfq.rfq_number} — {rfq.equipment_tag} {rfq.equipment_description} — TECHNICAL EVALUATION"

    ws["B2"] = rfq.project_name
    ws["B3"] = rfq.location
    ws["B4"] = rfq.owner
    ws["B5"] = rfq.engineer_buyer
    ws["B6"] = f"{rfq.rfq_number} — {rfq.equipment_tag} {rfq.equipment_description}"
    ws["B7"] = rfq.data_sheet_ref
    ws["B8"] = rfq.bid_tab_rev
    ws["B9"] = rfq.bid_tab_date or ""
    ws["B10"] = rfq.notes or ""

    # ------------------------------------------------------------------
    # 2. Vendor headers (row 12) and info rows 13-18
    # ------------------------------------------------------------------
    # Columns: D = Bidder 1, F = Bidder 2, H = Bidder 3
    vendor_cols = [4, 6, 8]  # D, F, H

    for i, col in enumerate(vendor_cols):
        if i < len(vendors):
            v = vendors[i]
            label = f"{v.vendor.name} ({v.vendor.revision or v.vendor.proposal_number})"
            ws.cell(row=12, column=col, value=label)
        else:
            ws.cell(row=12, column=col, value="")

    # Info rows
    info_map = [
        (13, lambda v: v.vendor.proposal_number),
        (14, lambda v: v.vendor.proposal_date or ""),
        (15, lambda v: v.vendor.sales_contact or ""),
        (16, lambda v: f"{v.vendor.phone or ''} / {v.vendor.email or ''}".strip(" /")),
        (17, lambda v: f"{v.vendor.shop_location or ''} ({v.vendor.fob or ''})".strip()),
        (18, lambda v: v.vendor.asme_stamp or ""),
    ]

    for row, getter in info_map:
        for i, col in enumerate(vendor_cols):
            if i < len(vendors):
                ws.cell(row=row, column=col, value=getter(vendors[i]))
            else:
                ws.cell(row=row, column=col, value="")

    # ------------------------------------------------------------------
    # 3. Pricing section (rows 22+)
    # Template has 5 data rows (22-26) + totals row 27.
    # We will use the existing rows and formulas. If more components are
    # needed later we can insert, but for the demo 5 is enough.
    # ------------------------------------------------------------------
    # Collect a stable ordered list of pricing items across vendors
    item_order = []
    seen = set()
    for v in vendors:
        for p in v.pricing_lines:
            key = p.item
            if key not in seen:
                seen.add(key)
                item_order.append(p)

    # Limit to the 5 pre-formatted rows in the template for now
    price_start = 22
    max_price_rows = 5

    for idx in range(max_price_rows):
        row = price_start + idx
        if idx < len(item_order):
            item = item_order[idx]
            ws.cell(row=row, column=1, value=item.item)          # Component
            ws.cell(row=row, column=2, value=item.description)   # Description

            for i, col in enumerate(vendor_cols):
                # Bidder price column
                price_col = col
                amount = 0.0
                if i < len(vendors):
                    for p in vendors[i].pricing_lines:
                        if p.item == item.item:
                            amount = p.amount_usd
                            break
                ws.cell(row=row, column=price_col, value=amount if amount else 0)

            # Re-apply / keep the 20% formulas for Vendor Data Costs
            # C = 20% of D, E = 20% of F, G = 20% of H
            ws.cell(row=row, column=3, value=f"=D{row}*0.2")
            ws.cell(row=row, column=5, value=f"=F{row}*0.2")
            ws.cell(row=row, column=7, value=f"=H{row}*0.2")
        else:
            # Clear unused rows
            for c in range(1, 10):
                if c in (3, 5, 7):
                    ws.cell(row=row, column=c, value=f"={get_column_letter(c+1)}{row}*0.2")
                else:
                    ws.cell(row=row, column=c, value=0 if c in (4, 6, 8) else "")

    # Totals row (27) – keep / fix SUM formulas so they cover the 5 rows
    ws["A27"] = "NORMALIZED TOTALS"
    ws["C27"] = "=SUM(C22:C26)"
    ws["D27"] = "=SUM(D22:D26)"
    ws["E27"] = "=SUM(E22:E26)"
    ws["F27"] = "=SUM(F22:F26)"
    ws["G27"] = "=SUM(G22:G26)"
    ws["H27"] = "=SUM(H22:H26)"

    # ------------------------------------------------------------------
    # 4. Commercial Terms (rows 31-38)
    # ------------------------------------------------------------------
    # Header row 30 — use real vendor names + rev (same style as Pricing)
    for i, col in enumerate(vendor_cols):
        if i < len(vendors):
            v = vendors[i]
            label = f"{v.vendor.name} ({v.vendor.revision or v.vendor.proposal_number})"
            ws.cell(row=30, column=col, value=label)
        else:
            ws.cell(row=30, column=col, value="")

    term_rows = [
        (31, "Delivery", lambda v: v.commercial.delivery_weeks),
        (32, "Pricing Validity", lambda v: v.commercial.pricing_validity),
        (33, "Payment Terms", lambda v: v.commercial.payment_terms),
        (34, "Warranty", lambda v: v.commercial.warranty),
        (35, "Cancellation", lambda v: v.commercial.cancellation_schedule),
        (36, "Taxes", lambda v: v.commercial.taxes),
        (37, "Tariff / Govt Action", lambda v: v.commercial.tariff_disclaimer),
        (38, "Storage", lambda v: v.commercial.storage),
    ]

    for row, _label, getter in term_rows:
        for i, col in enumerate(vendor_cols):
            if i < len(vendors):
                ws.cell(row=row, column=col, value=getter(vendors[i]) or "")
            else:
                ws.cell(row=row, column=col, value="")

    # ------------------------------------------------------------------
    # 5. Technical Evaluation sheet
    # ------------------------------------------------------------------
    # Update bidder headers on row 4
    tech_vendor_cols = [4, 5, 6]  # D, E, F
    for i, col in enumerate(tech_vendor_cols):
        if i < len(vendors):
            v = vendors[i]
            ws2.cell(row=4, column=col, value=f"{v.vendor.name} ({v.vendor.revision or ''})")
        else:
            ws2.cell(row=4, column=col, value="")

    # Fill technical parameters under the correct category headers.
    # Template category header rows (merged) and the data rows that follow them:
    # GENERAL = row 5, data 6-8
    # DESIGN DATA = row 9, data 10-19
    # MATERIALS OF CONSTRUCTION = row 20, data 21-30
    # EXTERNAL & INTERNAL ATTACHMENTS = row 31, data 32-38
    # NOZZLE SCHEDULE = row 39, data 40-54
    # INTERNALS = row 55, data 56+
    category_row_map = {
        "GENERAL": list(range(6, 9)),
        "DESIGN DATA": list(range(10, 20)),
        "MATERIALS OF CONSTRUCTION": list(range(21, 31)),
        "MATERIALS": list(range(21, 31)),
        "EXTERNAL & INTERNAL ATTACHMENTS": list(range(32, 39)),
        "ATTACHMENTS": list(range(32, 39)),
        "NOZZLE SCHEDULE": list(range(40, 55)),
        "INTERNALS": list(range(56, 58)),
        "HYDROTEST": list(range(10, 20)),  # fold into DESIGN DATA
    }

    if vendors:
        from collections import defaultdict
        # Group master parameters by category (use first vendor as master order)
        by_cat = defaultdict(list)
        for t in vendors[0].technical:
            by_cat[t.category].append(t)

        for cat, params in by_cat.items():
            rows = category_row_map.get(cat, [])
            for idx, t in enumerate(params):
                if idx >= len(rows):
                    break
                row = rows[idx]
                ws2.cell(row=row, column=1, value=t.parameter)
                ws2.cell(row=row, column=2, value=t.units or "—")
                ws2.cell(row=row, column=3, value=t.data_sheet_required)

                for i, v in enumerate(vendors):
                    match = next((x for x in v.technical if x.parameter == t.parameter and x.category == t.category), None)
                    if match:
                        ws2.cell(row=row, column=4 + i, value=match.vendor_offer)
                        if i == 0:
                            ws2.cell(row=row, column=7, value=match.compliance.value)
                    else:
                        ws2.cell(row=row, column=4 + i, value="—")

                if t.notes:
                    ws2.cell(row=row, column=8, value=t.notes)

    # ------------------------------------------------------------------
    # Save to bytes
    # ------------------------------------------------------------------
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
