"""
Excel export that closely matches the professional Bid Tab layout
used by Crosstrails (22510-M001-Bid-Tab-Rev 1 style).
"""

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from app.models import BidTab, calculate_apples_to_apples


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
thin = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
section_fill = PatternFill("solid", fgColor="D6DCE4")
section_font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
label_font = Font(name='Calibri', bold=True, size=10)
normal_font = Font(name='Calibri', size=10)
title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
currency_format = '"$"#,##0'
total_fill = PatternFill("solid", fgColor="E2EFDA")
total_font = Font(name='Calibri', bold=True, size=10)
dev_fill = PatternFill("solid", fgColor="FCE4D6")  # light orange for deviations
ok_fill = PatternFill("solid", fgColor="C6EFCE")   # light green


def _set_col_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_header_block(ws, bt: BidTab):
    """Write the project header that matches the original Bid Tab."""
    rfq = bt.rfq

    # Title row
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{rfq.project_number} KEYSTONE C2 PURIFICATION — {rfq.equipment_tag} TOWER ({rfq.equipment_description.upper()}) — BID TAB"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # Project info block
    info = [
        ("Project:", rfq.project_name),
        ("Location:", rfq.location),
        ("Owner:", rfq.owner),
        ("Engineer / Buyer:", rfq.engineer_buyer),
        ("RFQ Package:", f"{rfq.rfq_number} — Tower ({rfq.equipment_tag} Demethanizer Column)"),
        ("Data Sheet:", rfq.data_sheet_ref),
        ("Bid Tab Rev:", rfq.bid_tab_rev),
        ("Bid Tab Date:", rfq.bid_tab_date or datetime.now().strftime("%d %B %Y")),
        ("Revisions:", rfq.notes or ""),
    ]

    for i, (label, value) in enumerate(info, start=2):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = label_font
        ws.merge_cells(f'B{i}:H{i}')
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = normal_font


def build_classic_excel(bt: BidTab) -> bytes:
    """
    Build a Classic 2-sheet Bid Tab that closely follows the layout
    of 22510-M001-Bid-Tab-Rev 1.xlsx
    """
    wb = Workbook()

    # =====================================================================
    # SHEET 1: Commercial Summary
    # =====================================================================
    ws = wb.active
    ws.title = "Commercial Summary"

    _write_header_block(ws, bt)

    # ---- Vendor Info section ----
    start_row = 12
    ws[f'A{start_row}'] = "Vendor Info"
    ws[f'A{start_row}'].font = section_font
    ws[f'A{start_row}'].fill = section_fill
    ws.merge_cells(f'A{start_row}:H{start_row}')

    # Column headers for vendors
    header_row = start_row + 1
    ws[f'B{header_row}'] = "Field"
    ws[f'B{header_row}'].font = header_font
    ws[f'B{header_row}'].fill = header_fill

    vendors = bt.vendors
    # Place vendors starting at column D, F, H style (or simply sequential)
    vendor_cols = ['D', 'F', 'H']  # supports up to 3 cleanly; extend if needed
    for i, v in enumerate(vendors):
        col = vendor_cols[i] if i < len(vendor_cols) else get_column_letter(4 + i * 2)
        cell = ws[f'{col}{header_row}']
        cell.value = f"{v.vendor.name} ({v.vendor.revision or ''})".strip()
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center')

    ws[f'H{header_row}'] = "Notes"
    ws[f'H{header_row}'].font = header_font
    ws[f'H{header_row}'].fill = header_fill

    # Vendor info rows
    info_fields = [
        ("Proposal / Quote #", lambda v: v.vendor.proposal_number),
        ("Proposal Date", lambda v: v.vendor.proposal_date or "—"),
        ("Sales Contact", lambda v: v.vendor.sales_contact or "—"),
        ("Phone / Email", lambda v: f"{v.vendor.phone or ''} / {v.vendor.email or ''}".strip(" /")),
        ("Shop Location / FOB", lambda v: f"{v.vendor.shop_location or ''} ({v.vendor.fob or ''})".strip()),
        ("ASME / NB Stamp", lambda v: v.vendor.asme_stamp or "—"),
    ]

    for offset, (label, getter) in enumerate(info_fields):
        r = header_row + 1 + offset
        ws[f'B{r}'] = label
        ws[f'B{r}'].font = label_font
        for i, v in enumerate(vendors):
            col = vendor_cols[i] if i < len(vendor_cols) else get_column_letter(4 + i * 2)
            ws[f'{col}{r}'] = getter(v)
            ws[f'{col}{r}'].font = normal_font
            ws[f'{col}{r}'].alignment = Alignment(wrap_text=True)

    # ---- Pricing Summary ----
    price_start = header_row + len(info_fields) + 2
    ws[f'A{price_start}'] = "PRICING SUMMARY (USD)"
    ws[f'A{price_start}'].font = section_font
    ws[f'A{price_start}'].fill = section_fill
    ws.merge_cells(f'A{price_start}:H{price_start}')

    # Pricing header
    ph = price_start + 1
    headers = ["Item", "Description", "Absolute (USD)", "BWFS (USD)", "GLEX (USD)", "Notes"]
    # Dynamic based on actual vendors
    price_headers = ["Item", "Description"]
    for v in vendors:
        price_headers.append(f"{v.vendor.name.split()[0]} (USD)")
    price_headers.append("Notes")

    for col_idx, h in enumerate(price_headers, start=1):
        cell = ws.cell(row=ph, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Collect common pricing items
    # We use a simple ordered list of known items
    item_order = ["Base Vessel", "Downcomer Piping", "Ladders & Platforms", "Fall Arrest Devices", "Field Hydrotest"]

    price_row = ph + 1
    for item_name in item_order:
        ws.cell(row=price_row, column=1, value=item_name).font = label_font

        # Description from first vendor that has it
        desc = ""
        for v in vendors:
            for p in v.pricing_lines:
                if item_name.lower() in p.item.lower() or item_name.lower() in p.description.lower():
                    desc = p.description
                    break
            if desc:
                break
        ws.cell(row=price_row, column=2, value=desc).font = normal_font
        ws.cell(row=price_row, column=2).alignment = Alignment(wrap_text=True)

        for i, v in enumerate(vendors):
            amount = 0.0
            is_opt = False
            for p in v.pricing_lines:
                if item_name.lower() in p.item.lower() or (item_name == "Ladders & Platforms" and "ladder" in p.item.lower()):
                    amount = p.amount_usd
                    is_opt = p.is_optional
                    break
            cell = ws.cell(row=price_row, column=3 + i, value=amount if amount else None)
            cell.number_format = currency_format
            cell.font = normal_font
            if is_opt:
                cell.font = Font(name='Calibri', size=10, italic=True, color='808080')

        price_row += 1

    # Apples-to-apples total row
    total_row = price_row
    ws.cell(row=total_row, column=1, value="TOTAL — Apples-to-Apples").font = total_font
    ws.cell(row=total_row, column=2, value=bt.apples_to_apples_notes or "Vessel + downcomers + base ladders").font = normal_font
    for i, v in enumerate(vendors):
        total = calculate_apples_to_apples(v)
        cell = ws.cell(row=total_row, column=3 + i, value=total)
        cell.number_format = currency_format
        cell.font = total_font
        cell.fill = total_fill

    # ---- Commercial Terms ----
    terms_start = total_row + 2
    ws[f'A{terms_start}'] = "COMMERCIAL TERMS — VENDOR COMPARISON"
    ws[f'A{terms_start}'].font = section_font
    ws[f'A{terms_start}'].fill = section_fill
    ws.merge_cells(f'A{terms_start}:H{terms_start}')

    th = terms_start + 1
    term_headers = ["Term", "Description"] + [v.vendor.name.split()[0] for v in vendors] + ["Notes / Risk"]
    for col_idx, h in enumerate(term_headers, start=1):
        cell = ws.cell(row=th, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    term_fields = [
        ("Delivery", "Equipment delivery (weeks)", lambda v: v.commercial.delivery_weeks),
        ("Pricing Validity", "Quote validity period", lambda v: v.commercial.pricing_validity),
        ("Payment Terms", "Progress payment milestones", lambda v: v.commercial.payment_terms),
        ("Warranty", "Mechanical warranty", lambda v: v.commercial.warranty),
        ("Cancellation", "Cancellation schedule", lambda v: v.commercial.cancellation_schedule),
        ("Taxes", "Sales / use / interstate tax", lambda v: v.commercial.taxes),
        ("Tariff / Govt Action", "Tariff & regulatory disclaimer", lambda v: v.commercial.tariff_disclaimer),
        ("Storage", "Equipment storage post-completion", lambda v: v.commercial.storage),
    ]

    for offset, (term, desc, getter) in enumerate(term_fields):
        r = th + 1 + offset
        ws.cell(row=r, column=1, value=term).font = label_font
        ws.cell(row=r, column=2, value=desc).font = normal_font
        for i, v in enumerate(vendors):
            val = getter(v) or "—"
            cell = ws.cell(row=r, column=3 + i, value=val)
            cell.font = normal_font
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths for Commercial sheet
    _set_col_widths(ws, {
        'A': 22, 'B': 42, 'C': 18, 'D': 22, 'E': 18, 'F': 22, 'G': 18, 'H': 28
    })

    # =====================================================================
    # SHEET 2: Technical Comparison
    # =====================================================================
    ws2 = wb.create_sheet("Technical Comparison")

    # Title
    ws2.merge_cells('A1:G1')
    ws2['A1'] = f"{bt.rfq.equipment_tag} DEMETHANIZER COLUMN — TECHNICAL COMPARISON vs DATA SHEET {bt.rfq.data_sheet_ref.split('(')[0].strip()}"
    ws2['A1'].font = title_font
    ws2.row_dimensions[1].height = 22

    ws2.merge_cells('A2:G2')
    ws2['A2'] = f"{bt.rfq.bid_tab_rev}: Compliance legend: OK = matches DS; Dev = deviates; Clarify = not addressed / requires vendor clarification."
    ws2['A2'].font = Font(name='Calibri', size=9, italic=True)

    # Headers
    tech_headers = ["Parameter", "Units", "Data Sheet — Required"] + [v.vendor.name.split()[0] for v in vendors] + ["Compliance"]
    for col_idx, h in enumerate(tech_headers, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Group by category
    current_category = None
    row = 5

    # Collect ordered unique parameters preserving category order
    seen = set()
    ordered_params = []
    for v in bt.vendors:
        for t in v.technical:
            key = (t.category, t.parameter)
            if key not in seen:
                seen.add(key)
                ordered_params.append(t)

    for param in ordered_params:
        # Category header
        if param.category != current_category:
            current_category = param.category
            ws2.cell(row=row, column=1, value=current_category).font = section_font
            ws2.cell(row=row, column=1).fill = section_fill
            for c in range(2, len(tech_headers) + 1):
                ws2.cell(row=row, column=c).fill = section_fill
            row += 1

        ws2.cell(row=row, column=1, value=param.parameter).font = label_font
        ws2.cell(row=row, column=2, value=param.units or "—").font = normal_font
        ws2.cell(row=row, column=3, value=param.data_sheet_required).font = normal_font
        ws2.cell(row=row, column=3).alignment = Alignment(wrap_text=True)

        compliances = []
        for i, v in enumerate(vendors):
            match = next((t for t in v.technical if t.parameter == param.parameter and t.category == param.category), None)
            if match:
                cell = ws2.cell(row=row, column=4 + i, value=match.vendor_offer)
                cell.font = normal_font
                cell.alignment = Alignment(wrap_text=True)
                compliances.append(match.compliance.value)
                if match.compliance.value == "Dev":
                    cell.fill = dev_fill
            else:
                ws2.cell(row=row, column=4 + i, value="—").font = normal_font

        # Overall compliance (simplified)
        overall = "OK"
        if "Dev" in compliances:
            overall = "Dev"
        elif "Clarify" in compliances:
            overall = "Clarify"
        cell = ws2.cell(row=row, column=4 + len(vendors), value=overall)
        cell.font = label_font
        if overall == "OK":
            cell.fill = ok_fill
        elif overall == "Dev":
            cell.fill = dev_fill

        row += 1

    _set_col_widths(ws2, {
        'A': 28, 'B': 12, 'C': 45, 'D': 35, 'E': 35, 'F': 20, 'G': 12
    })

    # Freeze panes
    ws.freeze_panes = 'A12'
    ws2.freeze_panes = 'A5'

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
